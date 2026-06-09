from __future__ import annotations

import importlib.util
import json
import math
import sys
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(r"E:\professional management systems\stockflow")
MODEL_PATH = ROOT / "scripts" / "generate_bakery_recipe_book.py"
OUTPUT_JSON = ROOT / "outputs" / "bakery_recipe_book_data.json"


def load_model_module():
    spec = importlib.util.spec_from_file_location("bakery_recipe_model", MODEL_PATH)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Could not load model module at {MODEL_PATH}")
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


def clean_value(value):
    if isinstance(value, set):
        return sorted(value)
    if isinstance(value, Path):
        return str(value)
    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return None
        return value
    if hasattr(value, "__dict__"):
        return clean_value(value.__dict__)
    if isinstance(value, list):
        return [clean_value(item) for item in value]
    if isinstance(value, dict):
        return {str(key): clean_value(item) for key, item in value.items()}
    return value


def main() -> None:
    model = load_model_module()
    wb_values = load_workbook(model.SOURCE_WORKBOOK, data_only=True)
    wb_formulas = load_workbook(model.SOURCE_WORKBOOK, data_only=False)

    recipes = [
        recipe
        for sheet in wb_values.sheetnames
        if sheet not in model.EXCLUDED_RECIPE_SHEETS
        for recipe in [model.parse_recipe(sheet, wb_values, wb_formulas)]
        if recipe
    ]
    skus = model.build_skus(recipes)
    raw_materials = model.parse_raw_materials(wb_values)

    formula_errors = []
    for ws in wb_values.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("#"):
                    formula_errors.append(f"{ws.title}!{cell.coordinate}={cell.value}")

    ingredient_usage = {}
    for recipe in recipes:
        for ingredient in recipe["ingredients"]:
            group = ingredient["group"]
            usage = ingredient_usage.setdefault(group, {"recipes": set(), "qty": 0.0, "cost": 0.0})
            usage["recipes"].add(recipe["sheet"])
            usage["qty"] += ingredient.get("qty_0") or 0
            usage["cost"] += ingredient.get("cost_0") or 0

    recipe_actions = {}
    for recipe in recipes:
        recipe_skus = [sku for sku in skus if sku["recipe"] == recipe["sheet"]]
        action, reason = model.recommendation(recipe, recipe_skus)
        recipe_actions[recipe["sheet"]] = {"action": action, "reason": reason}

    data = {
        "source_workbook": str(model.SOURCE_WORKBOOK),
        "source_workbook_name": model.SOURCE_WORKBOOK.name,
        "target_gross_margin": model.TARGET_GROSS_MARGIN,
        "target_food_cost": model.TARGET_FOOD_COST,
        "benchmark_sources": model.BENCHMARK_SOURCES,
        "recipes": recipes,
        "skus": skus,
        "raw_materials": raw_materials,
        "ingredient_usage": ingredient_usage,
        "formula_errors": formula_errors,
        "recipe_actions": recipe_actions,
    }

    OUTPUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_JSON.write_text(json.dumps(clean_value(data), ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {OUTPUT_JSON}")
    print(f"Recipes={len(recipes)} SKUs={len(skus)} FormulaErrors={len(formula_errors)}")


if __name__ == "__main__":
    main()
