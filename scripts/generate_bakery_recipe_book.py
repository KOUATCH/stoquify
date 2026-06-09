from __future__ import annotations

import datetime as dt
import math
import re
from dataclasses import dataclass
from pathlib import Path
from statistics import median
from typing import Any

from openpyxl import load_workbook


SOURCE_WORKBOOK = Path(r"E:\retail management systems\Bakery recipes   melto 2026.xlsx")
LOCK_FILE = Path(r"E:\retail management systems\~$Bakery recipes   melto 2026.xlsx")
OUTPUT = Path(r"E:\professional management systems\stockflow\bakery_comprehensive_recipe_book.md")

RAW_MATERIAL_SHEET = " COST rRAW MATERIAL"
EXCLUDED_RECIPE_SHEETS = {RAW_MATERIAL_SHEET, "Millitary Records"}
TARGET_GROSS_MARGIN = 0.65
TARGET_FOOD_COST = 1 - TARGET_GROSS_MARGIN

BENCHMARK_SOURCES = [
    {
        "label": "KitchenCost recipe-costing guide",
        "url": "https://kitchencost.app/en/blog/recipe-costing/",
        "use": "Recipe costing formulas, cost-per-serving, menu price = cost / target food cost, and the warning that food cost excludes labor/overhead.",
    },
    {
        "label": "Sage restaurant prime-cost benchmark",
        "url": "https://www.sage.com/en-us/blog/restaurant-prime-costs/",
        "use": "Prime cost, food cost, and labor cost benchmark ranges for operating context.",
    },
    {
        "label": "RestaurantOwner prime-cost reference",
        "url": "https://www.restaurantowner.com/prime.pdf",
        "use": "Prime cost threshold context and why food cost alone is not enough.",
    },
]


@dataclass
class Variant:
    name: str
    qty_col: int
    cost_col: int
    mass_total: float | None = None
    cost_total: float | None = None
    mass_src: str = ""
    cost_src: str = ""


@dataclass
class OutputSku:
    row: int
    label: str | None
    price: float
    weight: float
    price_src: str
    weight_src: str


def normalize(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip().lower())


def is_number(value: Any) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool) and math.isfinite(value)


def fmt_money(value: float | None) -> str:
    if value is None or not is_number(value):
        return "n/a"
    if abs(value) >= 1000:
        return f"{value:,.0f}"
    return f"{value:,.2f}"


def fmt_number(value: float | None) -> str:
    if value is None or not is_number(value):
        return "n/a"
    if abs(value) >= 1000:
        return f"{value:,.0f}"
    if abs(value - round(value)) < 1e-9:
        return f"{value:.0f}"
    return f"{value:.2f}"


def fmt_percent(value: float | None) -> str:
    if value is None or not is_number(value):
        return "n/a"
    return f"{value * 100:.1f}%"


def sheet_cell(sheet: str, row: int, col: int, wb_values) -> str:
    return f"{sheet}!{wb_values[sheet].cell(row, col).coordinate}"


def recipe_category(sheet_name: str) -> str:
    name = normalize(sheet_name)
    if any(token in name for token in ["short bread", "danisa", "delice", "zebree", "okinawa", "melto"]):
        return "biscuit/sweet"
    if any(token in name for token in ["cake", "gateau", "gato", "marbre"]):
        return "cake/gateau"
    if "buns" in name:
        return "buns"
    if any(token in name for token in ["brioche", "pain", "baguette", "bageutte", "bread", "banh", "ngala", "yummy", "universal"]):
        return "bread"
    if any(token in name for token in ["donut", "pancake", "chinchin", "sugar balls"]):
        return "fried/snack"
    if any(token in name for token in ["fish pie", "galette", "croissant"]):
        return "pastry/savory"
    if "ice cream" in name:
        return "ice cream"
    return "other"


def ingredient_group(name: str) -> str:
    value = normalize(name).replace("cold water", "water").strip()
    aliases = {
        "eggs": "egg",
        "egg carton": "egg",
        "egg tray": "egg",
        "butter 2": "butter",
        "frying oil": "oil",
        " frying oil": "oil",
        "conc milk": "milk",
        "milk tantalizer": "milk",
        "powder milk": "milk",
        "icing sugar": "sugar",
        "nut meg": "nutmeg",
        "nutmeg": "nutmeg",
        "flour bir": "flour",
    }
    return aliases.get(value, value)


def standard_method(category: str) -> list[str]:
    methods = {
        "bread": [
            "Scale ingredients exactly from the recipe card; keep flour, yeast, salt, sugar, fat, and improver separate until mixing.",
            "Mix dry ingredients, add water/liquids gradually, then develop dough until smooth and elastic.",
            "Bulk ferment until dough shows visible rise; divide by target portion weight, round, rest, shape, proof, bake, cool fully, then package.",
        ],
        "buns": [
            "Scale and mix as enriched dough; add fat after initial hydration if dough development is weak.",
            "Divide to target portion weight, round tightly, proof consistently, bake/fry according to house standard, cool, then package.",
            "Track piece count against theoretical yield after cooling, not before, because shrink and breakage affect saleable units.",
        ],
        "cake/gateau": [
            "Cream fat and sugar or mix according to house cake method; add eggs/liquids slowly to avoid splitting.",
            "Fold dry ingredients gently, portion by weight, bake until set, cool fully before cutting or packing.",
            "Track batter weight, baked weight, trim, and final saleable pieces for yield control.",
        ],
        "fried/snack": [
            "Mix dough/batter to consistent hydration, rest where required, portion or cut by target sale weight.",
            "Fry in controlled oil; drain fully before weighing and packing so oil pickup is visible in yield records.",
            "Track oil usage and discard schedule separately because current workbook treats oil as an ingredient but not a process loss.",
        ],
        "pastry/savory": [
            "Prepare dough and filling separately; weigh filling per piece to control cost and consistency.",
            "Seal, proof/rest where required, bake/fry, cool, then count only intact saleable pieces.",
            "Track filling waste and broken/leaking pieces because savory items have higher hidden loss risk.",
        ],
        "biscuit/sweet": [
            "Mix dough until uniform, avoid overworking after flour addition, portion or sheet to consistent thickness.",
            "Bake/fry according to house standard, cool fully, then pack by weight or count.",
            "Track broken pieces, trim, and underweight packs as shrink.",
        ],
        "ice cream": [
            "Confirm recipe units before production; current workbook does not provide freezing loss or churn yield.",
            "Batch, freeze/churn according to house process, then measure final saleable volume/weight.",
            "Track container loss and temperature-related shrink separately.",
        ],
    }
    return methods.get(category, ["Use the house standard method and add exact process steps, temperatures, times, and yield checkpoints."])


def find_item_header(ws_formulas) -> tuple[int, int] | None:
    for row in ws_formulas.iter_rows():
        for cell in row:
            if normalize(cell.value) == "item":
                return cell.row, cell.column
    return None


def is_ingredient_name(value: Any) -> bool:
    name = normalize(value)
    if not name:
        return False
    stop_words = [
        "total",
        "totals",
        "bread type",
        "price",
        "chosen weight",
        "farine",
        "farinre",
        "choix",
        "profit",
        "mass",
        "cost",
        "sales price",
    ]
    return not any(name == word or name.startswith(word) for word in stop_words)


def parse_recipe(sheet: str, wb_values, wb_formulas) -> dict[str, Any] | None:
    ws = wb_values[sheet]
    wsf = wb_formulas[sheet]
    header = find_item_header(wsf)
    if not header:
        return None

    header_row, item_col = header
    variants: list[Variant] = []
    for col in range(item_col + 1, ws.max_column):
        label = normalize(wsf.cell(header_row, col).value or ws.cell(header_row, col).value)
        next_label = normalize(wsf.cell(header_row, col + 1).value or ws.cell(header_row, col + 1).value)
        looks_like_quantity = (
            "qtty" in label
            or "qty" in label
            or "mass" in label
            or (re.search(r"\b\d+(?:\.\d+)?\s*kg\b", label) and "cost" not in label)
            or (label.endswith("kg") and "cost" not in label)
        )
        if label and looks_like_quantity and "cost" not in label and "weight" not in label and "cost" in next_label:
            variants.append(Variant(str(wsf.cell(header_row, col).value), col, col + 1))

    ingredient_rows: list[int] = []
    for row in range(header_row + 1, ws.max_row + 1):
        if is_ingredient_name(ws.cell(row, item_col).value):
            ingredient_rows.append(row)
        elif ingredient_rows:
            break

    ingredients: list[dict[str, Any]] = []
    for row in ingredient_rows:
        ingredient = {
            "name": str(ws.cell(row, item_col).value).strip(),
            "row": row,
            "group": ingredient_group(str(ws.cell(row, item_col).value)),
        }
        for idx, variant in enumerate(variants):
            qty = ws.cell(row, variant.qty_col).value
            cost = ws.cell(row, variant.cost_col).value
            ingredient[f"qty_{idx}"] = qty if is_number(qty) else None
            ingredient[f"cost_{idx}"] = cost if is_number(cost) else None
            ingredient[f"qty_src_{idx}"] = sheet_cell(sheet, row, variant.qty_col, wb_values)
            ingredient[f"cost_src_{idx}"] = sheet_cell(sheet, row, variant.cost_col, wb_values)
        ingredients.append(ingredient)

    for idx, variant in enumerate(variants):
        mass = sum((ingredient.get(f"qty_{idx}") or 0) for ingredient in ingredients)
        cost = sum((ingredient.get(f"cost_{idx}") or 0) for ingredient in ingredients)
        variant.mass_total = mass if mass else None
        variant.cost_total = cost if cost else None
        variant.mass_src = " + ".join(
            ingredient[f"qty_src_{idx}"] for ingredient in ingredients if ingredient.get(f"qty_{idx}") is not None
        )
        variant.cost_src = " + ".join(
            ingredient[f"cost_src_{idx}"] for ingredient in ingredients if ingredient.get(f"cost_{idx}") is not None
        )

    price_headers: list[tuple[int, int, int]] = []
    for row in range(header_row + 1, ws.max_row + 1):
        for col in range(1, ws.max_column):
            label = normalize(ws.cell(row, col).value)
            next_label = normalize(ws.cell(row, col + 1).value)
            if (
                label == "price"
                and (
                    next_label == "weight"
                    or (
                        row + 1 <= ws.max_row
                        and is_number(ws.cell(row + 1, col).value)
                        and is_number(ws.cell(row + 1, col + 1).value)
                    )
                )
            ) or (label == "bread type" and next_label == "weight"):
                price_headers.append((row, col, col + 1))
            if (
                label == "bread type"
                and row + 2 <= ws.max_row
                and normalize(ws.cell(row + 1, col).value) == "price"
                and is_number(ws.cell(row + 2, col).value)
                and is_number(ws.cell(row + 2, col + 1).value)
            ):
                price_headers.append((row + 1, col, col + 1))

    outputs: list[OutputSku] = []
    seen: set[tuple[int, int, int]] = set()
    for header_row, price_col, weight_col in price_headers:
        blanks = 0
        for row in range(header_row + 1, min(ws.max_row + 1, header_row + 12)):
            price = ws.cell(row, price_col).value
            weight = ws.cell(row, weight_col).value
            if is_number(price) and is_number(weight) and price > 0 and weight > 0:
                key = (row, price_col, weight_col)
                if key in seen:
                    continue
                seen.add(key)
                label: str | None = None
                for col in range(1, price_col):
                    value = ws.cell(row, col).value
                    if isinstance(value, str) and value.strip().strip("*"):
                        label = value.strip()
                        break
                outputs.append(
                    OutputSku(
                        row=row,
                        label=label,
                        price=float(price),
                        weight=float(weight),
                        price_src=sheet_cell(sheet, row, price_col, wb_values),
                        weight_src=sheet_cell(sheet, row, weight_col, wb_values),
                    )
                )
                blanks = 0
            else:
                if seen:
                    blanks += 1
                if blanks >= 3:
                    break

    base = variants[0] if variants else None
    top_ingredients = []
    if base and base.cost_total:
        top_ingredients = sorted(
            [ingredient for ingredient in ingredients if ingredient.get("cost_0") is not None],
            key=lambda item: item.get("cost_0") or 0,
            reverse=True,
        )

    return {
        "sheet": sheet,
        "category": recipe_category(sheet),
        "header": sheet_cell(sheet, header_row, item_col, wb_values),
        "variants": variants,
        "ingredients": ingredients,
        "outputs": outputs,
        "base_mass": base.mass_total if base else None,
        "base_cost": base.cost_total if base else None,
        "top_ingredients": top_ingredients,
        "top2_share": (
            sum((ingredient.get("cost_0") or 0) for ingredient in top_ingredients[:2]) / base.cost_total
            if base and base.cost_total and top_ingredients
            else None
        ),
    }


def parse_raw_materials(wb_values) -> list[dict[str, Any]]:
    if RAW_MATERIAL_SHEET not in wb_values.sheetnames:
        return []
    ws = wb_values[RAW_MATERIAL_SHEET]
    materials = []
    for row in range(1, ws.max_row + 1):
        name = ws.cell(row, 2).value
        cost = ws.cell(row, 3).value
        weight = ws.cell(row, 4).value
        if isinstance(name, str) and name.strip():
            if is_number(cost) or is_number(weight):
                unit = (float(cost) / float(weight)) if is_number(cost) and is_number(weight) and weight else None
                materials.append(
                    {
                        "name": name.strip(),
                        "cost": float(cost) if is_number(cost) else None,
                        "weight": float(weight) if is_number(weight) else None,
                        "unit": unit,
                        "source": f"{RAW_MATERIAL_SHEET}!B{row}:D{row}",
                    }
                )
    return materials


def build_skus(recipes: list[dict[str, Any]]) -> list[dict[str, Any]]:
    skus: list[dict[str, Any]] = []
    for recipe in recipes:
        base_cost = recipe.get("base_cost")
        base_mass = recipe.get("base_mass")
        for output in recipe["outputs"]:
            cogs = base_cost * output.weight / base_mass if base_cost and base_mass else None
            yield_units = base_mass / output.weight if base_mass and output.weight else None
            gross_profit = output.price - cogs if cogs is not None else None
            gross_margin = gross_profit / output.price if gross_profit is not None and output.price else None
            food_cost = cogs / output.price if cogs is not None and output.price else None
            target_price = cogs / TARGET_FOOD_COST if cogs is not None else None
            price_gap = target_price - output.price if target_price is not None else None
            max_profitable_weight = output.price * TARGET_FOOD_COST * base_mass / base_cost if base_cost and base_mass else None
            break_even_units = base_cost / output.price if base_cost and output.price else None
            top = recipe["top_ingredients"][0] if recipe["top_ingredients"] else None
            top_cost_unit = ((top.get("cost_0") or 0) * output.weight / base_mass) if top and base_mass else None
            skus.append(
                {
                    "recipe": recipe["sheet"],
                    "category": recipe["category"],
                    "label": output.label,
                    "price": output.price,
                    "weight": output.weight,
                    "price_src": output.price_src,
                    "weight_src": output.weight_src,
                    "cogs": cogs,
                    "yield_units": yield_units,
                    "gross_profit": gross_profit,
                    "gross_margin": gross_margin,
                    "food_cost": food_cost,
                    "target_price": target_price,
                    "price_gap": price_gap,
                    "max_profitable_weight": max_profitable_weight,
                    "break_even_units": break_even_units,
                    "top_ingredient": top["name"] if top else None,
                    "top_cost_unit": top_cost_unit,
                    "margin_after_top_10": (
                        (output.price - (cogs + top_cost_unit * 0.10)) / output.price
                        if cogs is not None and top_cost_unit is not None and output.price
                        else None
                    ),
                    "margin_after_top_25": (
                        (output.price - (cogs + top_cost_unit * 0.25)) / output.price
                        if cogs is not None and top_cost_unit is not None and output.price
                        else None
                    ),
                }
            )
    return skus


def verdict(gross_margin: float | None) -> str:
    if gross_margin is None:
        return "⚪"
    if gross_margin < 0.50:
        return "🔴"
    if gross_margin < TARGET_GROSS_MARGIN:
        return "🟡"
    return "🟢"


def recommendation(recipe: dict[str, Any], recipe_skus: list[dict[str, Any]]) -> tuple[str, str]:
    if not recipe_skus:
        return "Retire pending audit", "No auditable sale price/portion row was parsed; do not keep in active costing until price/yield are fixed."
    worst = min(recipe_skus, key=lambda item: item["gross_margin"] if item["gross_margin"] is not None else -999)
    top2_share = recipe.get("top2_share")
    if worst["gross_margin"] is not None and worst["gross_margin"] < 0.35:
        if top2_share and top2_share > 0.60:
            return "Reformulate + reprice", f"Worst SKU is {fmt_percent(worst['gross_margin'])} GM and top-2 ingredients drive {fmt_percent(top2_share)} of base cost."
        return "Resize portion + reprice", f"Worst SKU is {fmt_percent(worst['gross_margin'])} GM; current portion is too large for the price."
    if worst["gross_margin"] is not None and worst["gross_margin"] < TARGET_GROSS_MARGIN:
        return "Reprice", f"Raise worst SKU to target price or reduce portion to {fmt_number(worst['max_profitable_weight'])}g for 65% GM."
    if top2_share and top2_share > 0.60:
        return "Renegotiate ingredient", f"Margin is acceptable but top-2 ingredients drive {fmt_percent(top2_share)} of base cost."
    return "Maintain + complete data", "Ingredient margin is acceptable; add labor, packaging, actual yield, and waste tracking."


def main() -> None:
    wb_values = load_workbook(SOURCE_WORKBOOK, data_only=True)
    wb_formulas = load_workbook(SOURCE_WORKBOOK, data_only=False)

    recipes = [
        recipe
        for sheet in wb_values.sheetnames
        if sheet not in EXCLUDED_RECIPE_SHEETS
        for recipe in [parse_recipe(sheet, wb_values, wb_formulas)]
        if recipe
    ]
    skus = build_skus(recipes)
    raw_materials = parse_raw_materials(wb_values)

    category_prices: dict[str, list[float]] = {}
    for sku in skus:
        category_prices.setdefault(sku["category"], []).append(sku["price"] / sku["weight"])
    category_medians = {category: median(values) for category, values in category_prices.items() if values}

    ingredient_usage: dict[str, dict[str, Any]] = {}
    for recipe in recipes:
        for ingredient in recipe["ingredients"]:
            group = ingredient["group"]
            usage = ingredient_usage.setdefault(group, {"recipes": set(), "qty": 0.0, "cost": 0.0})
            usage["recipes"].add(recipe["sheet"])
            usage["qty"] += ingredient.get("qty_0") or 0
            usage["cost"] += ingredient.get("cost_0") or 0

    formula_errors = []
    for ws in wb_values.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("#"):
                    formula_errors.append(f"{ws.title}!{cell.coordinate}={cell.value}")

    red = sum(1 for sku in skus if sku["gross_margin"] is not None and sku["gross_margin"] < 0.50)
    yellow = sum(1 for sku in skus if sku["gross_margin"] is not None and 0.50 <= sku["gross_margin"] < TARGET_GROSS_MARGIN)
    green = sum(1 for sku in skus if sku["gross_margin"] is not None and sku["gross_margin"] >= TARGET_GROSS_MARGIN)
    concentration = [recipe for recipe in recipes if recipe.get("top2_share") and recipe["top2_share"] > 0.60]

    lines: list[str] = []
    lines.append("# Melto Bakery Comprehensive Recipe Book and Profit Playbook")
    lines.append("")
    lines.append("## Source Register")
    lines.append("| Status | Source | Modified | Role |")
    lines.append("|---|---|---:|---|")
    source_modified = dt.datetime.fromtimestamp(SOURCE_WORKBOOK.stat().st_mtime).isoformat(timespec="seconds")
    lines.append(
        f"| Ingested | `{SOURCE_WORKBOOK}` | {source_modified} | Recipe/BOM sheets, raw ingredient costs, selling prices, portion weights, cached formulas. |"
    )
    if LOCK_FILE.exists():
        lock_modified = dt.datetime.fromtimestamp(LOCK_FILE.stat().st_mtime).isoformat(timespec="seconds")
        lines.append(f"| Excluded | `{LOCK_FILE}` | {lock_modified} | Excel temporary lock file; not a business-data source. |")
    for source in BENCHMARK_SOURCES:
        lines.append(f"| Benchmark | [{source['label']}]({source['url']}) | current web reference checked 2026-05-31 | {source['use']} |")
    lines.append("")
    lines.append("## How To Use This Book")
    lines.append("- Treat workbook cells as source-of-truth for ingredient quantities, costs, prices, and portion weights.")
    lines.append("- Treat any method, benchmark, target price, and profit action as a draft operating control until validated by production tests.")
    lines.append("- Target economics in this book use 35% ingredient food cost / 65% ingredient gross margin because labor, packaging, overhead, and waste are not in the workbook.")
    lines.append("- Contribution margin currently equals ingredient gross profit because variable labor, packaging, and overhead are missing.")
    lines.append("- Break-even units are ingredient-cost break-even per base batch only; daily break-even needs fixed costs and daily volume.")
    lines.append("")
    lines.append("## Portfolio Summary")
    lines.append("| Metric | Result | Source |")
    lines.append("|---|---:|---|")
    lines.append(f"| Recipe/BOM sheets parsed | {len(recipes)} | Workbook sheet list in `{SOURCE_WORKBOOK.name}` |")
    lines.append(f"| Saleable SKU/portion rows parsed | {len(skus)} | Price/weight tables in recipe sheets |")
    lines.append(f"| Red margin SKUs below 50% GM | {red} | Calculated from ingredient COGS and sale price cells |")
    lines.append(f"| Yellow margin SKUs at 50-65% GM | {yellow} | Calculated from ingredient COGS and sale price cells |")
    lines.append(f"| Green margin SKUs at or above 65% GM | {green} | Calculated from ingredient COGS and sale price cells |")
    lines.append(f"| Recipes with top-2 ingredient concentration above 60% | {len(concentration)} | Ingredient cost rows in recipe sheets |")
    lines.append(f"| Cached workbook formula errors found | {len(formula_errors)} | Workbook cached cell values |")
    lines.append("")

    lines.append("## Master Profitability Matrix")
    lines.append("| Recipe / SKU | Category | Price | Portion | Ingredient COGS | GM % | Food cost % | Target price @35% FC | Gap | Max weight at current price | Action | Source |")
    lines.append("|---|---|---:|---:|---:|---:|---:|---:|---:|---:|---|---|")
    for sku in sorted(skus, key=lambda item: (item["gross_margin"] if item["gross_margin"] is not None else 999, item["recipe"], item["weight"])):
        recipe = next(item for item in recipes if item["sheet"] == sku["recipe"])
        action, _reason = recommendation(recipe, [item for item in skus if item["recipe"] == recipe["sheet"]])
        lines.append(
            f"| {sku['recipe']} {fmt_number(sku['weight'])}g | {sku['category']} | {fmt_money(sku['price'])} | {fmt_number(sku['weight'])}g | {fmt_money(sku['cogs'])} | {verdict(sku['gross_margin'])} {fmt_percent(sku['gross_margin'])} | {fmt_percent(sku['food_cost'])} | {fmt_money(sku['target_price'])} | {fmt_money(sku['price_gap'])} | {fmt_number(sku['max_profitable_weight'])}g | {action} | `{sku['price_src']}`, `{sku['weight_src']}` + base ingredient rows |"
        )
    lines.append("")

    lines.append("## Raw Material Cost Index")
    lines.append("| Ingredient | Package cost | Package weight | Unit cost | Source |")
    lines.append("|---|---:|---:|---:|---|")
    for material in raw_materials:
        lines.append(
            f"| {material['name']} | {fmt_money(material['cost'])} | {fmt_number(material['weight'])} | {fmt_money(material['unit'])} | `{material['source']}` |"
        )
    lines.append("")

    lines.append("## Purchasing Consolidation Priorities")
    lines.append("| Ingredient group | Recipes using it | Base-batch spend | Base-batch quantity | Recommended control |")
    lines.append("|---|---:|---:|---:|---|")
    for group, usage in sorted(ingredient_usage.items(), key=lambda item: item[1]["cost"], reverse=True)[:20]:
        control = "Bid/renegotiate this week" if len(usage["recipes"]) >= 8 or usage["cost"] > 50000 else "Monitor and standardize pack size"
        lines.append(f"| {group} | {len(usage['recipes'])} | {fmt_money(usage['cost'])} | {fmt_number(usage['qty'])} | {control} |")
    lines.append("")

    lines.append("## Recipe Cards")
    for recipe in recipes:
        recipe_skus = [sku for sku in skus if sku["recipe"] == recipe["sheet"]]
        action, reason = recommendation(recipe, recipe_skus)
        base_variant = recipe["variants"][0] if recipe["variants"] else None
        top = recipe["top_ingredients"][0] if recipe["top_ingredients"] else None
        lines.append(f"### {recipe['sheet']}")
        lines.append(f"| Field | Value | Source / note |")
        lines.append("|---|---|---|")
        lines.append(f"| Category | {recipe['category']} | Inferred from recipe name |")
        lines.append(f"| Base batch mass | {fmt_number(recipe.get('base_mass'))} | `{base_variant.mass_src if base_variant else recipe['header']}` |")
        lines.append(f"| Base ingredient cost | {fmt_money(recipe.get('base_cost'))} | `{base_variant.cost_src if base_variant else recipe['header']}` |")
        lines.append(f"| Current verdict | {action} | {reason} |")
        lines.append(f"| Top-2 ingredient concentration | {fmt_percent(recipe.get('top2_share'))} | Base ingredient cost table |")
        if top:
            lines.append(f"| Highest-cost ingredient | {top['name']} ({fmt_money(top.get('cost_0'))}) | `{top.get('cost_src_0')}` |")
        lines.append("| Missing data to complete COGS | labor minutes/rate; packaging; overhead; actual yield; waste/shrinkage; daily volume | Add these before treating total COGS as final |")
        lines.append("")

        lines.append("#### Ingredients")
        lines.append("| Ingredient | Qty | Ingredient cost | % of base cost | Cost-risk note | Source |")
        lines.append("|---|---:|---:|---:|---|---|")
        base_cost = recipe.get("base_cost")
        for ingredient in sorted(recipe["ingredients"], key=lambda item: item.get("cost_0") or 0, reverse=True):
            share = (ingredient.get("cost_0") or 0) / base_cost if base_cost else None
            risk = "High" if share and share >= 0.30 else ("Medium" if share and share >= 0.15 else "Normal")
            lines.append(
                f"| {ingredient['name']} | {fmt_number(ingredient.get('qty_0'))} | {fmt_money(ingredient.get('cost_0'))} | {fmt_percent(share)} | {risk} | `{ingredient.get('qty_src_0')}`, `{ingredient.get('cost_src_0')}` |"
            )
        lines.append("")

        lines.append("#### Saleable Portions and Profit Targets")
        lines.append("| Portion | Current price | Ingredient COGS | Gross profit | GM % | Food cost % | Target price @35% FC | Price lift needed | Max weight at current price | Break-even units / base batch | Source |")
        lines.append("|---:|---:|---:|---:|---:|---:|---:|---:|---:|---:|---|")
        if recipe_skus:
            for sku in recipe_skus:
                lines.append(
                    f"| {fmt_number(sku['weight'])}g | {fmt_money(sku['price'])} | {fmt_money(sku['cogs'])} | {fmt_money(sku['gross_profit'])} | {verdict(sku['gross_margin'])} {fmt_percent(sku['gross_margin'])} | {fmt_percent(sku['food_cost'])} | {fmt_money(sku['target_price'])} | {fmt_money(sku['price_gap'])} | {fmt_number(sku['max_profitable_weight'])}g | {fmt_number(sku['break_even_units'])} | `{sku['price_src']}`, `{sku['weight_src']}` |"
                )
        else:
            lines.append("| n/a | missing | n/a | n/a | n/a | n/a | n/a | n/a | n/a | n/a | No saleable price/portion row parsed |")
        lines.append("")

        lines.append("#### Ingredient Inflation Sensitivity")
        lines.append("| Portion | Key ingredient | GM after key +10% | GM after key +25% | Profit action if spike happens | Source |")
        lines.append("|---:|---|---:|---:|---|---|")
        if recipe_skus and top:
            for sku in recipe_skus:
                spike_action = "Raise price immediately" if sku["margin_after_top_25"] is not None and sku["margin_after_top_25"] < 0.50 else "Monitor / absorb short term"
                lines.append(
                    f"| {fmt_number(sku['weight'])}g | {top['name']} | {fmt_percent(sku['margin_after_top_10'])} | {fmt_percent(sku['margin_after_top_25'])} | {spike_action} | `{top.get('cost_src_0')}`, `{sku['price_src']}`, `{sku['weight_src']}` |"
                )
        else:
            lines.append("| n/a | missing | n/a | n/a | Complete ingredient and price rows | n/a |")
        lines.append("")

        lines.append("#### Standard Production Method")
        lines.append("⚠️ assumption: method steps are professional operating templates, not sourced from the workbook. Validate with your baker before production.")
        for step_number, step in enumerate(standard_method(recipe["category"]), start=1):
            lines.append(f"{step_number}. {step}")
        lines.append("")

        lines.append("#### Profit Improvement Checklist")
        checklist = [
            f"Set target selling price from the table above, or resize to the listed max weight at current price.",
            "Record actual good units after cooling/packing and compare to theoretical yield.",
            "Add labor minutes by step, packaging per unit, and overhead allocation per batch.",
            "Review top ingredient supplier price and minimum-order quantity.",
            "If formula errors exist in this sheet, repair workbook formulas before relying on workbook profit rows.",
        ]
        for item in checklist:
            lines.append(f"- {item}")
        lines.append("")

    lines.append("## Recipe Data Completion Template")
    lines.append("| Recipe | Labor minutes | Loaded labor rate | Packaging/unit | Overhead/batch | Actual good units | Waste units | Owner | Due date |")
    lines.append("|---|---:|---:|---:|---:|---:|---:|---|---|")
    for recipe in recipes:
        lines.append(f"| {recipe['sheet']} | missing | missing | missing | missing | missing | missing | Production + Finance | next costing cycle |")
    lines.append("")

    lines.append("## Formula and Data Hygiene Flags")
    lines.append("| Flag | Detail | Action |")
    lines.append("|---|---|---|")
    if formula_errors:
        lines.append(f"| Workbook formula errors | {', '.join(formula_errors[:20])} | Repair `#REF!` cells before using workbook profit rows. |")
    else:
        lines.append("| Workbook formula errors | None found in cached values | Continue formula checks during each update. |")
    duplicate_names = [recipe["sheet"] for recipe in recipes if normalize(recipe["sheet"]) in {"sheet3"}]
    if duplicate_names:
        lines.append(f"| Possible duplicate/template sheet | {', '.join(duplicate_names)} | Rename or archive after confirming whether it is an active recipe. |")
    lines.append("| Non-operational text in raw material sheet | Credential/payment-looking strings were not printed in this book. | Remove secrets or unrelated text from the workbook. |")
    lines.append("")

    lines.append("## Assumptions")
    lines.append("| Assumption | Why it was used | What to replace it with |")
    lines.append("|---|---|---|")
    lines.append(f"| ⚠️ Target food cost is {fmt_percent(TARGET_FOOD_COST)} / target ingredient GM is {fmt_percent(TARGET_GROSS_MARGIN)}. | Bakery/restaurant costing references support using food-cost targets, but your labor and overhead are missing. | Your actual prime-cost model by department and sales channel. |")
    lines.append("| ⚠️ Base recipe uses the first quantity/cost pair after the `item` column. | Workbook contains multiple scaled-batch columns with inconsistent labels/formulas. | A selected-batch flag per recipe. |")
    lines.append("| ⚠️ Portion weights are treated as grams or workbook-native weight units. | Workbook labels say `weight` but not a formal unit. | Confirm unit of measure for each recipe. |")
    lines.append("| ⚠️ Target price ignores demand elasticity. | No sales-volume history or demand curve is present. | POS sales by SKU before/after price changes. |")
    lines.append("| ⚠️ Contribution margin equals gross profit. | Labor, packaging, and variable overhead are absent. | Labor minutes, loaded wage, packaging/unit, utility and rent allocation. |")

    OUTPUT.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"Wrote {OUTPUT}")
    print(f"Recipes={len(recipes)} SKUs={len(skus)} Red={red} Yellow={yellow} Green={green} FormulaErrors={len(formula_errors)}")


if __name__ == "__main__":
    main()
